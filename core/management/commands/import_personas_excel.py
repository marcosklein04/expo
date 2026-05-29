from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path
import unicodedata

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import IntegrityError

from core.models import Empresa, Persona
from core.services import normalizar_codigo_empresa, normalizar_dni, normalizar_texto

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise ImportError("openpyxl no esta instalado") from exc


COLUMN_ALIASES = {
    "dni": "dni",
    "documento": "dni",
    "nro dni": "dni",
    "nombre y apellido": "nombre_apellido",
    "nombre apellido": "nombre_apellido",
    "apellido y nombre": "nombre_apellido",
    "nombre": "nombre",
    "apellido": "apellido",
    "concesionario": "concesionario",
    "credencial": "credencial",
    "tipo de vianda": "tipo_vianda",
    "menu": "tipo_vianda",
    "menú": "tipo_vianda",
    "vianda": "tipo_vianda",
    "puede invitar": "puede_invitar",
    "invitar": "puede_invitar",
    "habilita invitados": "puede_invitar",
    "invitaciones": "puede_invitar",
}
REQUIRED_FIELDS = {"dni", "nombre_apellido"}
LAYOUT_AUTO = "auto"
LAYOUT_DNI = "dni"
LAYOUT_VALTRA_FENDT = "valtra_fendt"
LAYOUT_VALTRA_FENDT_DNI = "valtra_fendt_dni"
LAYOUT_MASSEY_DODECAEDRO = "massey_dodecaedro"

SPLIT_NAME_REQUIRED_FIELDS = {"dni", "nombre", "apellido"}
CONCESIONARIO_HEADER_TOKENS = {"concesionario", "concesionarios"}
VIANDA_HEADER_TOKENS = {
    "clasico": "CLASICO",
    "clasica": "CLASICO",
    "vegetariano": "VEGETARIANO",
    "vegetariana": "VEGETARIANO",
    "celiaco": "CELIACO",
    "celiaca": "CELIACO",
    "sin tacc": "CELIACO",
    "sintacc": "CELIACO",
    "sin gluten": "CELIACO",
}


def _normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return " ".join(text.replace("\n", " ").split())


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _cell_to_bool(value: object) -> bool:
    text = _cell_to_text(value).strip().lower()
    if not text:
        return False
    return text in {"1", "si", "sí", "s", "true", "x", "yes", "y"}


def _cell_to_vianda(value: object) -> str:
    text = _cell_to_text(value).strip().lower()
    if text in {"vegetariano", "veg", "vegetariana"}:
        return Persona.VIANDA_VEGETARIANO
    if text in {"celiaco", "celiaca", "sin tacc", "sintacc", "sin gluten"}:
        return Persona.VIANDA_CELIACO
    return Persona.VIANDA_CLASICO


def _special_guest_names() -> set[str]:
    raw_names = getattr(settings, "KIOSK_SPECIAL_GUEST_NAMES", []) or []
    return {normalizar_texto(name) for name in raw_names if str(name or "").strip()}


def _should_enable_guests(nombre_apellido: str) -> bool:
    return normalizar_texto(nombre_apellido) in _special_guest_names()


def _find_header_row(worksheet) -> tuple[int, dict[str, int]]:
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        header_map: dict[str, int] = {}
        for index, raw_cell in enumerate(row):
            normalized = _normalize_header(raw_cell)
            if not normalized:
                continue
            mapped = COLUMN_ALIASES.get(normalized)
            if mapped and mapped not in header_map:
                header_map[mapped] = index

        if REQUIRED_FIELDS.issubset(header_map.keys()):
            return row_index, header_map

    raise CommandError(
        "No se encontro una fila de cabecera valida. Debe incluir al menos DNI y Nombre y Apellido."
    )


def _detect_layout(worksheet) -> str:
    try:
        _find_header_row(worksheet)
        return LAYOUT_DNI
    except CommandError:
        pass

    for row in worksheet.iter_rows(min_row=1, max_row=6, values_only=True):
        normalized = [_normalize_header(value) for value in row]
        if "credencial" in normalized and "nombre" in normalized and "apellido" in normalized:
            return LAYOUT_VALTRA_FENDT

    try:
        _find_split_name_header_row(worksheet)
        return LAYOUT_VALTRA_FENDT_DNI
    except CommandError:
        pass

    raise CommandError("No se pudo detectar automaticamente el layout del Excel.")


def _resolve_empresa(*, empresa_code: str | None, empresa_name: str | None) -> tuple[Empresa, bool]:
    raw_empresa_code = empresa_code or getattr(settings, "DEFAULT_EMPRESA_CODE", "DEFAULT")
    codigo = normalizar_codigo_empresa(str(raw_empresa_code)) or "DEFAULT"
    nombre = str(empresa_name or codigo).strip() or codigo

    empresa, created = Empresa.objects.get_or_create(
        codigo=codigo,
        defaults={"nombre": nombre, "activo": True},
    )
    if empresa.nombre != nombre:
        empresa.nombre = nombre
        empresa.save(update_fields=["nombre", "actualizado_en"])
    if not empresa.activo:
        empresa.activo = True
        empresa.save(update_fields=["activo", "actualizado_en"])
    return empresa, created


def _save_persona_with_dni(
    *,
    empresa: Empresa,
    dni: str,
    defaults: dict[str, object],
) -> tuple[Persona, bool, bool]:
    obj = Persona.objects.filter(empresa=empresa, dni=dni).first()
    was_created = False
    relinked_document = False
    nombre_apellido = str(defaults["nombre_apellido"])
    concesionario = str(defaults.get("concesionario", "") or "")
    credencial = str(defaults.get("credencial", "") or "")

    if obj is None and any(ch.isalpha() for ch in dni):
        candidate_qs = Persona.objects.filter(
            empresa=empresa,
            nombre_apellido=nombre_apellido,
            activo=True,
        )
        if concesionario:
            candidate_qs = candidate_qs.filter(concesionario=concesionario)
        if credencial:
            candidate_qs = candidate_qs.filter(credencial=credencial)

        if candidate_qs.count() == 1:
            candidate = candidate_qs.first()
            if candidate and candidate.dni.isdigit():
                candidate.dni = dni
                for field, value in defaults.items():
                    setattr(candidate, field, value)
                try:
                    candidate.save()
                    obj = candidate
                    relinked_document = True
                except IntegrityError:
                    obj = None

    if obj is None:
        obj = Persona.objects.create(empresa=empresa, dni=dni, **defaults)
        was_created = True
    else:
        update_fields = ["activo", "actualizado_en"]
        for field, value in defaults.items():
            setattr(obj, field, value)
            update_fields.append(field)
        obj.save(update_fields=sorted(set(update_fields)))

    return obj, was_created, relinked_document


def _iter_header_rows(worksheet, *, header_row: int, header_map: dict[str, int]) -> Iterable[dict[str, object]]:
    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        raw_dni = _cell_to_text(row[header_map["dni"]])
        dni = normalizar_dni(raw_dni)
        nombre_apellido = _cell_to_text(row[header_map["nombre_apellido"]])
        if not dni or not nombre_apellido:
            yield {"skip": True}
            continue

        concesionario = _cell_to_text(row[header_map["concesionario"]]) if "concesionario" in header_map else ""
        credencial = _cell_to_text(row[header_map["credencial"]]) if "credencial" in header_map else ""
        puede_invitar = (
            _cell_to_bool(row[header_map["puede_invitar"]])
            if "puede_invitar" in header_map
            else _should_enable_guests(nombre_apellido)
        )
        tipo_vianda = (
            _cell_to_vianda(row[header_map["tipo_vianda"]])
            if "tipo_vianda" in header_map
            else Persona.VIANDA_CLASICO
        )
        yield {
            "dni": dni,
            "nombre_apellido": nombre_apellido,
            "concesionario": concesionario,
            "credencial": credencial,
            "tipo_vianda": tipo_vianda,
            "puede_invitar": puede_invitar,
            "activo": True,
        }


def _iter_valtra_fendt_rows(worksheet) -> Iterable[dict[str, object]]:
    current_concesionario = ""
    current_credencial = ""

    for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if row_idx < 6:
            continue

        concesionario = _cell_to_text(row[1] if len(row) > 1 else "")
        credencial = _cell_to_text(row[3] if len(row) > 3 else "")
        nombre = _cell_to_text(row[4] if len(row) > 4 else "")
        apellido = _cell_to_text(row[5] if len(row) > 5 else "")
        if concesionario:
            current_concesionario = concesionario
        if credencial:
            current_credencial = credencial

        nombre_apellido = " ".join(part for part in [nombre, apellido] if part).strip()
        if not nombre_apellido:
            continue

        dias = [
            _cell_to_text(row[index] if len(row) > index else "")
            for index in (8, 9, 10, 11, 12)
        ]
        activo = any(value not in {"", "0"} for value in dias)

        yield {
            "row_idx": row_idx,
            "nombre_apellido": nombre_apellido,
            "concesionario": current_concesionario,
            "credencial": current_credencial,
            "tipo_vianda": _cell_to_vianda(row[7] if len(row) > 7 else ""),
            "puede_invitar": _should_enable_guests(nombre_apellido),
            "activo": activo,
        }


def _match_persona_by_name(
    *,
    empresa: Empresa,
    nombre_apellido: str,
    concesionario: str,
    credencial: str,
) -> Persona | None:
    normalized_name = normalizar_texto(nombre_apellido)
    candidates = [
        persona
        for persona in Persona.objects.filter(empresa=empresa)
        if normalizar_texto(persona.nombre_apellido) == normalized_name
    ]
    if len(candidates) == 1:
        return candidates[0]

    if credencial:
        normalized_credencial = normalizar_texto(credencial)
        candidates = [
            persona
            for persona in candidates
            if normalizar_texto(persona.credencial) == normalized_credencial
        ]
    if len(candidates) == 1:
        return candidates[0]

    if concesionario:
        normalized_concesionario = normalizar_texto(concesionario)
        candidates = [
            persona
            for persona in candidates
            if normalizar_texto(persona.concesionario) == normalized_concesionario
        ]
    if len(candidates) == 1:
        return candidates[0]

    return None


def _find_split_name_header_row(worksheet) -> tuple[int, dict[str, int], int | None]:
    """Detecta fila de headers con columnas separadas Nombre, Apellido, DNI.

    Devuelve (fila_header, mapeo_columnas, columna_concesionario_o_None).
    El detector escanea hasta 50 filas; en ese rango identifica:
    - la fila con los headers DNI + Nombre + Apellido,
    - opcionalmente la columna del concesionario (si arriba o al costado dice
      'Concesionario(s)').
    """
    concesionario_col: int | None = None

    for row_index, row in enumerate(worksheet.iter_rows(min_row=1, max_row=50, values_only=True), start=1):
        header_map: dict[str, int] = {}
        for col_index, raw_cell in enumerate(row):
            normalized = _normalize_header(raw_cell)
            if not normalized:
                continue
            if normalized in CONCESIONARIO_HEADER_TOKENS and concesionario_col is None:
                concesionario_col = col_index
                continue
            mapped = COLUMN_ALIASES.get(normalized)
            if mapped and mapped not in header_map:
                header_map[mapped] = col_index

        if SPLIT_NAME_REQUIRED_FIELDS.issubset(header_map.keys()):
            return row_index, header_map, concesionario_col

    raise CommandError(
        "No se encontro una fila de cabecera valida para el layout Valtra/Fendt con DNI. "
        "Se requieren columnas DNI, Nombre y Apellido por separado."
    )


def _iter_valtra_fendt_dni_rows(worksheet) -> Iterable[dict[str, object]]:
    header_row, header_map, concesionario_col = _find_split_name_header_row(worksheet)

    dni_col = header_map["dni"]
    nombre_col = header_map["nombre"]
    apellido_col = header_map["apellido"]
    tipo_vianda_col = header_map.get("tipo_vianda")
    credencial_col = header_map.get("credencial")

    current_concesionario = ""

    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        if concesionario_col is not None and len(row) > concesionario_col:
            raw_conc = _cell_to_text(row[concesionario_col])
            if raw_conc:
                current_concesionario = raw_conc

        raw_dni = _cell_to_text(row[dni_col]) if len(row) > dni_col else ""
        nombre = _cell_to_text(row[nombre_col]) if len(row) > nombre_col else ""
        apellido = _cell_to_text(row[apellido_col]) if len(row) > apellido_col else ""

        nombre_apellido = " ".join(part for part in [nombre, apellido] if part).strip()
        dni = normalizar_dni(raw_dni)

        if not dni or not nombre_apellido:
            yield {"skip": True}
            continue

        tipo_vianda = (
            _cell_to_vianda(row[tipo_vianda_col])
            if tipo_vianda_col is not None and len(row) > tipo_vianda_col
            else Persona.VIANDA_CLASICO
        )
        credencial = (
            _cell_to_text(row[credencial_col])
            if credencial_col is not None and len(row) > credencial_col
            else ""
        )

        yield {
            "dni": dni,
            "nombre_apellido": nombre_apellido,
            "concesionario": current_concesionario,
            "credencial": credencial,
            "tipo_vianda": tipo_vianda,
            "puede_invitar": _should_enable_guests(nombre_apellido),
            "activo": True,
        }


def _find_massey_header_row(
    worksheet,
) -> tuple[int, dict[str, int], int | None, dict[str, int]]:
    """Detecta fila de headers del layout Massey/Dodecaedro.

    Devuelve (fila_header, mapeo_columnas, columna_concesionario, columnas_vianda).
    Busca DNI + Nombre y Apellido (combinado) + columna Concesionario (mergeada)
    + tres columnas Clasico/Vegetariano/Celiaco con marca X.
    """
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=50, values_only=True), start=1
    ):
        header_map: dict[str, int] = {}
        concesionario_col: int | None = None
        vianda_cols: dict[str, int] = {}

        for col_index, raw_cell in enumerate(row):
            normalized = _normalize_header(raw_cell)
            if not normalized:
                continue
            if normalized in CONCESIONARIO_HEADER_TOKENS and concesionario_col is None:
                concesionario_col = col_index
                continue
            if normalized in VIANDA_HEADER_TOKENS:
                vianda_codigo = VIANDA_HEADER_TOKENS[normalized]
                if vianda_codigo not in vianda_cols:
                    vianda_cols[vianda_codigo] = col_index
                continue
            mapped = COLUMN_ALIASES.get(normalized)
            if mapped and mapped not in header_map:
                header_map[mapped] = col_index

        if REQUIRED_FIELDS.issubset(header_map.keys()):
            return row_index, header_map, concesionario_col, vianda_cols

    raise CommandError(
        "No se encontro fila de cabecera valida para el layout Massey/Dodecaedro. "
        "Se requieren columnas DNI y Nombre y Apellido."
    )


def _iter_massey_dodecaedro_rows(worksheet) -> Iterable[dict[str, object]]:
    header_row, header_map, concesionario_col, vianda_cols = _find_massey_header_row(worksheet)

    dni_col = header_map["dni"]
    nombre_col = header_map["nombre_apellido"]
    credencial_col = header_map.get("credencial")

    current_concesionario = ""

    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        if concesionario_col is not None and len(row) > concesionario_col:
            raw_conc = _cell_to_text(row[concesionario_col])
            if raw_conc:
                current_concesionario = raw_conc

        raw_dni = _cell_to_text(row[dni_col]) if len(row) > dni_col else ""
        nombre_apellido = _cell_to_text(row[nombre_col]) if len(row) > nombre_col else ""
        dni = normalizar_dni(raw_dni)

        if not dni or not nombre_apellido:
            yield {"skip": True}
            continue

        tipo_vianda = Persona.VIANDA_CLASICO
        for vianda_codigo, col in vianda_cols.items():
            if col >= len(row):
                continue
            cell = row[col]
            if cell is None:
                continue
            cell_text = str(cell).strip()
            if cell_text and cell_text != "0":
                tipo_vianda = vianda_codigo
                break

        credencial = (
            _cell_to_text(row[credencial_col])
            if credencial_col is not None and len(row) > credencial_col
            else ""
        )

        yield {
            "dni": dni,
            "nombre_apellido": nombre_apellido,
            "concesionario": current_concesionario,
            "credencial": credencial,
            "tipo_vianda": tipo_vianda,
            "puede_invitar": _should_enable_guests(nombre_apellido),
            "activo": True,
        }


class Command(BaseCommand):
    help = "Importa personas desde archivo Excel (XLSX) a la tabla Persona."

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str, help="Ruta del archivo XLSX")
        parser.add_argument(
            "--sheet",
            type=str,
            default=None,
            help="Nombre de hoja (si se omite usa la activa)",
        )
        parser.add_argument(
            "--empresa-code",
            type=str,
            default=None,
            help="Codigo de empresa destino (si se omite usa DEFAULT_EMPRESA_CODE).",
        )
        parser.add_argument(
            "--empresa-name",
            type=str,
            default=None,
            help="Nombre de empresa (solo se usa al crear la empresa).",
        )
        parser.add_argument(
            "--layout",
            type=str,
            choices=[
                LAYOUT_AUTO,
                LAYOUT_DNI,
                LAYOUT_VALTRA_FENDT,
                LAYOUT_VALTRA_FENDT_DNI,
                LAYOUT_MASSEY_DODECAEDRO,
            ],
            default=LAYOUT_AUTO,
            help="Fuerza el parser a usar un layout especifico.",
        )

    def handle(self, *args, **options):
        xlsx_path = Path(options["xlsx_path"])
        if not xlsx_path.exists():
            raise CommandError(f"No existe el archivo: {xlsx_path}")

        workbook = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
        worksheet = workbook[options["sheet"]] if options["sheet"] else workbook.active
        empresa, empresa_created = _resolve_empresa(
            empresa_code=options["empresa_code"],
            empresa_name=options["empresa_name"],
        )
        if empresa_created:
            self.stdout.write(
                self.style.SUCCESS(f"Empresa creada: {empresa.codigo} ({empresa.nombre})")
            )

        layout = options["layout"]
        if layout == LAYOUT_AUTO:
            layout = _detect_layout(worksheet)
            self.stdout.write(self.style.WARNING(f"Layout detectado: {layout}"))

        if layout == LAYOUT_DNI:
            self._import_layout_with_dni(empresa=empresa, worksheet=worksheet)
            return
        if layout == LAYOUT_VALTRA_FENDT:
            self._import_valtra_fendt_layout(empresa=empresa, worksheet=worksheet)
            return
        if layout == LAYOUT_VALTRA_FENDT_DNI:
            target_sheets = (
                [workbook[options["sheet"]]]
                if options["sheet"]
                else [workbook[name] for name in workbook.sheetnames]
            )
            self._import_valtra_fendt_dni_layout(
                empresa=empresa, worksheets=target_sheets
            )
            return
        if layout == LAYOUT_MASSEY_DODECAEDRO:
            self._import_massey_dodecaedro_layout(empresa=empresa, worksheet=worksheet)
            return

        raise CommandError(f"Layout no soportado: {layout}")

    def _import_layout_with_dni(self, *, empresa: Empresa, worksheet) -> None:
        header_row, header_map = _find_header_row(worksheet)
        created = 0
        updated = 0
        relinked = 0
        skipped_empty = 0

        for record in _iter_header_rows(worksheet, header_row=header_row, header_map=header_map):
            if record.get("skip"):
                skipped_empty += 1
                continue

            dni = str(record["dni"])
            defaults = {
                "nombre_apellido": record["nombre_apellido"],
                "concesionario": record["concesionario"],
                "credencial": record["credencial"],
                "tipo_vianda": record["tipo_vianda"],
                "puede_invitar": record["puede_invitar"],
                "activo": record["activo"],
            }
            obj, was_created, relinked_document = _save_persona_with_dni(
                empresa=empresa,
                dni=dni,
                defaults=defaults,
            )

            if was_created:
                created += 1
            elif relinked_document:
                relinked += 1
                updated += 1
            else:
                updated += 1

            self.stdout.write(
                f"OK [{empresa.codigo}] {obj.dni} | {obj.nombre_apellido} | "
                f"{obj.concesionario} | {obj.credencial} | vianda={obj.tipo_vianda} | "
                f"puede_invitar={obj.puede_invitar}"
            )

        self.stdout.write(
            self.style.SUCCESS(
                f"Importacion finalizada para empresa={empresa.codigo}. "
                f"creados={created} actualizados={updated} relinked={relinked} omitidos={skipped_empty}"
            )
        )

    def _import_valtra_fendt_layout(self, *, empresa: Empresa, worksheet) -> None:
        updated = 0
        skipped_unmatched = 0
        skipped_ambiguous = 0

        for record in _iter_valtra_fendt_rows(worksheet):
            persona = _match_persona_by_name(
                empresa=empresa,
                nombre_apellido=str(record["nombre_apellido"]),
                concesionario=str(record["concesionario"]),
                credencial=str(record["credencial"]),
            )
            if persona is None:
                normalized_name = normalizar_texto(str(record["nombre_apellido"]))
                duplicates = [
                    obj
                    for obj in Persona.objects.filter(empresa=empresa)
                    if normalizar_texto(obj.nombre_apellido) == normalized_name
                ]
                if duplicates:
                    skipped_ambiguous += 1
                    self.stdout.write(
                        self.style.WARNING(
                            f"SKIP [{empresa.codigo}] fila={record['row_idx']} nombre={record['nombre_apellido']} "
                            "coincidencia ambigua sin DNI."
                        )
                    )
                else:
                    skipped_unmatched += 1
                    self.stdout.write(
                        self.style.WARNING(
                            f"SKIP [{empresa.codigo}] fila={record['row_idx']} nombre={record['nombre_apellido']} "
                            "sin coincidencia previa por nombre. El Excel no trae DNI."
                        )
                    )
                continue

            persona.nombre_apellido = str(record["nombre_apellido"])
            persona.concesionario = str(record["concesionario"])
            persona.credencial = str(record["credencial"])
            persona.tipo_vianda = str(record["tipo_vianda"])
            persona.puede_invitar = bool(record["puede_invitar"])
            persona.activo = bool(record["activo"])
            persona.save(
                update_fields=[
                    "nombre_apellido",
                    "concesionario",
                    "credencial",
                    "tipo_vianda",
                    "puede_invitar",
                    "activo",
                    "actualizado_en",
                ]
            )
            updated += 1
            self.stdout.write(
                f"OK [{empresa.codigo}] {persona.dni} | {persona.nombre_apellido} | "
                f"{persona.concesionario} | {persona.credencial} | vianda={persona.tipo_vianda} | "
                f"puede_invitar={persona.puede_invitar}"
            )

        self.stdout.write(
            self.style.SUCCESS(
                f"Importacion finalizada para empresa={empresa.codigo}. "
                f"actualizados={updated} omitidos_sin_match={skipped_unmatched} "
                f"omitidos_ambiguos={skipped_ambiguous}"
            )
        )

    def _import_valtra_fendt_dni_layout(
        self, *, empresa: Empresa, worksheets: list
    ) -> None:
        created = 0
        updated = 0
        relinked = 0
        skipped_empty = 0

        for worksheet in worksheets:
            self.stdout.write(self.style.WARNING(f"Procesando hoja: {worksheet.title}"))
            try:
                records = list(_iter_valtra_fendt_dni_rows(worksheet))
            except CommandError as exc:
                self.stdout.write(
                    self.style.WARNING(f"SKIP hoja {worksheet.title}: {exc}")
                )
                continue

            for record in records:
                if record.get("skip"):
                    skipped_empty += 1
                    continue

                dni = str(record["dni"])
                defaults = {
                    "nombre_apellido": record["nombre_apellido"],
                    "concesionario": record["concesionario"],
                    "credencial": record["credencial"],
                    "tipo_vianda": record["tipo_vianda"],
                    "puede_invitar": record["puede_invitar"],
                    "activo": record["activo"],
                }
                obj, was_created, relinked_document = _save_persona_with_dni(
                    empresa=empresa,
                    dni=dni,
                    defaults=defaults,
                )

                if was_created:
                    created += 1
                elif relinked_document:
                    relinked += 1
                    updated += 1
                else:
                    updated += 1

                self.stdout.write(
                    f"OK [{empresa.codigo}/{worksheet.title}] {obj.dni} | "
                    f"{obj.nombre_apellido} | {obj.concesionario} | "
                    f"vianda={obj.tipo_vianda} | puede_invitar={obj.puede_invitar}"
                )

        self.stdout.write(
            self.style.SUCCESS(
                f"Importacion finalizada para empresa={empresa.codigo}. "
                f"creados={created} actualizados={updated} relinked={relinked} "
                f"omitidos={skipped_empty}"
            )
        )

    def _import_massey_dodecaedro_layout(self, *, empresa: Empresa, worksheet) -> None:
        created = 0
        updated = 0
        relinked = 0
        skipped_empty = 0

        for record in _iter_massey_dodecaedro_rows(worksheet):
            if record.get("skip"):
                skipped_empty += 1
                continue

            dni = str(record["dni"])
            defaults = {
                "nombre_apellido": record["nombre_apellido"],
                "concesionario": record["concesionario"],
                "credencial": record["credencial"],
                "tipo_vianda": record["tipo_vianda"],
                "puede_invitar": record["puede_invitar"],
                "activo": record["activo"],
            }
            obj, was_created, relinked_document = _save_persona_with_dni(
                empresa=empresa,
                dni=dni,
                defaults=defaults,
            )

            if was_created:
                created += 1
            elif relinked_document:
                relinked += 1
                updated += 1
            else:
                updated += 1

            self.stdout.write(
                f"OK [{empresa.codigo}] {obj.dni} | {obj.nombre_apellido} | "
                f"{obj.concesionario} | vianda={obj.tipo_vianda} | "
                f"puede_invitar={obj.puede_invitar}"
            )

        self.stdout.write(
            self.style.SUCCESS(
                f"Importacion finalizada para empresa={empresa.codigo}. "
                f"creados={created} actualizados={updated} relinked={relinked} "
                f"omitidos={skipped_empty}"
            )
        )
